import os
import zipfile
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import pandas
from utils import path_resources, path_tmp, path_tmp_zip_file, path_resources_zip_file


def test_create_arch():
    os.chdir(path_resources)
    with zipfile.ZipFile('file.zip', 'w') as new_zip:
        new_zip.write('file.txt')
        new_zip.write('file.xls')
        new_zip.write('file.xlsx')
        new_zip.write('file.pdf')
    if not os.path.isdir(path_tmp):
        os.mkdir(path_tmp)
    if os.path.exists(path_tmp_zip_file):
        os.remove(path_tmp_zip_file)
    os.rename(path_resources_zip_file, path_tmp_zip_file)


def test_txt():
    with zipfile.ZipFile(path_tmp_zip_file) as zip_file:
        with zip_file.open('file.txt') as txt_f:
            assert txt_f.read().decode('utf-8') == 'Hello world'
            txt_file_size = zip_file.infolist()[0].file_size
            assert txt_file_size == 11


def test_xls():
    with zipfile.ZipFile(path_tmp_zip_file, 'r') as zip_file:
        with zip_file.open("file.xls") as xls_file:
            dt = pandas.read_excel(xls_file).head()
            assert 'Country' in dt
            xls_file_size = zip_file.infolist()[1].file_size
            assert xls_file_size == 8704


def test_xlsx_abc():
    with zipfile.ZipFile(path_tmp_zip_file) as zip_file:
        with zip_file.open("file.xlsx") as xlsx_data:
            workbook = load_workbook(xlsx_data)
            sheet = workbook.active
            sheet = sheet.cell(row=7, column=7).value
            assert sheet == '21/05/2015'
            xlsx_file_size = zip_file.infolist()[2].file_size
            assert xlsx_file_size == 7360


def test_pdf():
    with zipfile.ZipFile(path_tmp_zip_file) as zip_file:
        with zip_file.open('file.pdf') as pdf_file:
            pdf_data = PdfReader(pdf_file)
            assert "Python Testing with pytest\n" in pdf_data.pages[1].extract_text()
            number_of_page = len(pdf_data.pages)
            assert number_of_page == 256
            pdf_file_size = zip_file.infolist()[3].file_size
            assert pdf_file_size == 3035139
